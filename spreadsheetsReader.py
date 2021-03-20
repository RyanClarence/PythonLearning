import openpyxl


def list_company_with_inventory(file_name):
    ws = file_name["Sheet1"]
    product_per_supplier = {}
    for prod_row in range(2, ws.max_row + 1):
        supplier_name = ws.cell(prod_row, 4).value

        if supplier_name in product_per_supplier:
            current_num_products = product_per_supplier.get(supplier_name)
            product_per_supplier[supplier_name] = current_num_products + 1
        else:
            product_per_supplier[supplier_name] = 1
    print("Company Name\tNumber of Inventory")
    for key in product_per_supplier.keys():
        print("{}       {}".format(key, product_per_supplier.get(key)))


def total_product_cost(file_name):
    worksheet = file_name["Sheet1"]
    product_total_cost = {}

    for prod_row in range(2, worksheet.max_row + 1):
        price_per_product = worksheet.cell(prod_row, 3).value
        product_in_stock = worksheet.cell(prod_row, 2).value
        supplier_name = worksheet.cell(prod_row, 4).value
        product_name = worksheet.cell(prod_row, 1).value

        total_cost = price_per_product * product_in_stock
        worksheet["E1"] = "Total Price"
        worksheet[f"E{prod_row}"] = total_cost
        product_total_cost[product_name] = [supplier_name, product_name, product_in_stock, price_per_product, total_cost]

    file_name.save("inventoryFinal.xlsx")
    print("company_name\tproduct_name\tproduct_in_stock\tprice_per_product\ttotal_cost")
    for key in product_total_cost:
        print(product_total_cost[key])


if __name__ == "__main__":
    inv_file = openpyxl.load_workbook("inventory.xlsx")
    list_company_with_inventory(inv_file)
    total_product_cost(inv_file)
